
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'Rank'
ws['B1'] = 'Title'
ws['C1'] = 'Gross'
ws['D1'] = 'Theaters'
ws['E1'] = 'Avg Gross per Theater'


movie_rows = soup.findAll('tr')

for x in range(1,6):
    td = movie_rows[x].findAll('td')
    rank = td[0].text
    title = td[1].text
    gross = int(td[5].text.replace('$', '').replace(',', ''))
    theaters = int(td[6].text.replace(',' ,''))
    
    avg = gross/theaters

    ws['A1' + str(x+1)] = rank
    ws['B1' + str(x+1)] = title
    ws['C1' + str(x+1)] = gross
    ws['D1' + str(x+1)] = theaters
    ws['E1' + str(x+1)] = avg

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 35

header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

for cell in ws['D:D']:
    cell.number_format = '#,##0'

for cell in ws['C:C']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['E:E']:
    cell.number_format = u'"$ "#,##0.00'

wb.save('BoxOfficeReport.xlsx')