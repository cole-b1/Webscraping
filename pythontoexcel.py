import openpyxl as xl
from openpyxl.styles import Font

#create a new excel workbook
wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

#create a new worksheet
wb.create_sheet(index = 1, title = 'Second Sheet')

#write content to a cell
ws['A1'] = 'Invoice'

fontobj = Font(name='Times New Roman', size = 24, bold = True)

ws['A1'].font = fontobj

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws['B2'] = 450
ws['B3'] = 225.50
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = fontobj
  
ws.column_dimensions['A'].width = 25

ws.merge_cells('A1:B1')

ws['B8'] = '=SUM(B2:B7)'


#Produce Report Assignment
write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

rowcounter = 1
for row in read_ws.iter_rows():
    name = row[0].value
    cost = row[1].value
    amt_sold = row[2].value
    total = row[3].value

    write_sheet.cell(rowcounter, 1).value = name
    write_sheet.cell(rowcounter, 2).value = cost
    write_sheet.cell(rowcounter, 3).value = amt_sold
    write_sheet.cell(rowcounter, 4).value = total

    rowcounter += 1

summary_row = rowcounter + 2

write_sheet['B' + str(summary_row)] = 'Total'
write_sheet['B' + str(summary_row)].font = fontobj

write_sheet['C' + str(summary_row)] = '=SUM(C2:C' + str(rowcounter) + ')'
write_sheet['D' + str(summary_row)] = '=SUM(D2:D' + str(rowcounter) + ')'

summary_row += 1

write_sheet['B' + str(summary_row)] = 'Average'
write_sheet['B' + str(summary_row)].font = fontobj

write_sheet['C' + str(summary_row)] = '=Average(C2:C' + str(rowcounter) + ')'
write_sheet['D' + str(summary_row)] = '=Average(D2:D' + str(rowcounter) + ')'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 16
write_sheet.column_dimensions['C'].width = 16
write_sheet.column_dimensions['D'].width = 16

for cell in write_sheet['C:C']:
    cell.number_format = '#,##0'

for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'





#save the workbook
wb.save('PythonToExcel.xlsx')